import os
import shutil
import json
from datetime import datetime
from threading import Lock
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Side, Border, Color, PatternFill

root_path = os.path.dirname(__file__)


# Only for BenGoFrame Report
# This report will not create html by code
# Just generate test data details and modify to index.html in template folder
# Please do not delete or modify the fields in template folder
# testData.json is the copy file to display all test results.
class ReportCreator:
    def __init__(self, suite_name, grid_address, log_level, thread_count, database,
                 re_run,
                 output_path, start_time=""):
        """
        :param suite_name: Required
        :param grid_address: Required
        :param log_level: Required
        :param thread_count: Required
        :param database: Required
        :param re_run: Required
        :param output_path: Required
        :param start_time: string and Format (%Y-%m-%d %H:%M:%S).
        Will generate the start time when you create the reporter object if you don't input the start_time
        Up to you to set the start time. It should be at Hook @before_all/@before_all_test
        """
        self.output_path = output_path
        self.suite_name = suite_name
        self.grid_address = grid_address
        self.log_level = log_level
        self.thread_count = thread_count
        self.database = database
        self.re_run = re_run
        if not start_time:
            self.start_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            self.start_time = start_time
        self.end_time = ""
        self.take_times = ""
        self.testDetails = []
        self.testAll = {}
        self.config = self._set_config()
        self.appender = {}
        self.lock = Lock()

    def set_appender(self, name, appender):
        self.lock.acquire()
        try:
            self.appender[name] = appender
            self.lock.release()
        except Exception as e:
            self.lock.release()
            raise e

    def get_appender(self, name):
        return self.appender[name]

    def _set_config(self):
        config = {
            "testSuite": self.suite_name,
            "gridAddress": self.grid_address,
            "logLevel": self.log_level,
            "threadCount": self.thread_count,
            "database": self.database,
            "reRun": self.re_run,
            "startTime": self.start_time
        }
        return config

    def add_test_details(self, test_details):
        self.lock.acquire()
        try:
            self.testDetails.append(test_details)
            self.lock.release()
        except Exception as e:
            self.lock.release()
            raise e

    def completed(self, add_excel=False, screenshot_folder=""):
        time_format = '%Y-%m-%d %H:%M:%S'
        try:
            if not self.testDetails:
                raise Exception("You need create TestAppender and log test steps")
            # Reporter will log the end time when user call completed
            # Up to you to set the end time.
            # At below way is set the end time at Hook @after_all
            self.end_time = datetime.now().strftime(time_format)
            self.config['endTime'] = self.end_time
            take_times = str(
                datetime.strptime(self.end_time, time_format) - datetime.strptime(
                    self.start_time, time_format))
            group_times = take_times.split(":")
            self.config[
                'takeTimes'] = f'{group_times[0]} hs {group_times[1]} mins {group_times[2]} ss'
            # To build the summary details
            self._summary()
            # To build the feature summary details
            self._feature_summary()
            self.testAll['testDetails'] = self.testDetails
            self.testAll['testConfig'] = self.config
            # Copy the template report
            self._create_report_to()
            if not screenshot_folder:
                print(
                    "Your report will not display the screenshot "
                    "if you didn't package your screenshot folder")
            else:
                # Copy the input screenshot folder to report path src/static/img
                shutil.copytree(f'{screenshot_folder}', f'{self.output_path}/static/img',
                                dirs_exist_ok=True)
            # Set the json into the html
            self._modify_html()
            if add_excel:
                obj_excel = ExcelCreator(f"{self.output_path}/testData.json",
                                         f"{self.output_path}/TestReport.xlsx")
                obj_excel.save()
        except Exception as e:
            raise Exception(
                f"Create report failed. "
                f"Please make sure you had appended the test steps: {str(e)}")

    def _create_report_to(self):
        try:
            template_path = os.path.join(root_path, "template")
            index_path = os.path.join(root_path, "template/index.html")
            static_path = os.path.join(root_path, "template/static")
            if not os.path.exists(template_path):
                raise Exception(
                    "Please check if your template folder exist in src folder")
            elif not os.path.exists(index_path):
                raise Exception(
                    "Please check if your index.html exist in src/template folder")
            elif not os.path.exists(static_path):
                raise Exception(
                    "Please check if your static folder exist in src/template folder")
            shutil.copytree(template_path, self.output_path, dirs_exist_ok=True)
        except Exception as e:
            raise e

    def _summary(self):
        try:
            total_cases = len(self.testDetails)
            total_passed = 0
            total_failed = 0
            total_steps = 0
            total_passed_steps = 0
            total_failed_steps = 0
            for detail in self.testDetails:
                steps = detail['steps']
                total_steps += len(steps)
                if detail['status'] == 'passed':
                    total_passed += 1
                else:
                    total_failed += 1
                for step in steps:
                    if step['stepStatus'] == 'passed':
                        total_passed_steps += 1
                    else:
                        total_failed_steps += 1
            cases_chart = {'totalCases': total_cases, "passed": total_passed,
                           "failed": total_failed}
            steps_chart = {'totalSteps': total_steps, "passed": total_passed_steps,
                           "failed": total_failed_steps}
            self.testAll['casesChart'] = cases_chart
            self.testAll['stepsChart'] = steps_chart
        except Exception as e:
            raise e

    def _feature_summary(self):
        try:
            features = {}
            for detail in self.testDetails:
                summary = {"total": 0, "passed": 0, "failed": 0}
                feature_name = detail['featureName']
                feature_status = detail['status']
                if feature_name not in features:
                    if feature_status == 'passed':
                        summary['passed'] = 1
                    else:
                        summary['failed'] = 1
                    summary['total'] = 1
                    features[feature_name] = summary
                else:
                    if feature_status == 'passed':
                        features[feature_name]['passed'] += 1
                    else:
                        features[feature_name]['failed'] += 1
                    features[feature_name]['total'] += 1
            feature_summary = []
            for key, value in features.items():
                value['featureName'] = key
                value['passRate'] = f'{value["passed"] / value["total"] * 100:.2f}%'
                value["total"] = str(value['total'])
                value["passed"] = str(value['passed'])
                value["failed"] = str(value['failed'])
                feature_summary.append(value)
            self.testAll['featureSummary'] = feature_summary
        except Exception as e:
            raise e

    def _modify_html(self):
        try:
            # To make sure json always string
            if self.testAll is not str:
                self.testAll = str(self.testAll)
            test_data_json = json.dumps(self.testAll, indent=2)
            with open(f'{self.output_path}/testData.json', mode='w') as f:
                f.write(test_data_json)
                f.close()
            soup = BeautifulSoup(open(f'{self.output_path}/index.html', mode='r'), 'lxml')
            with open(f'{self.output_path}/index.html', mode='w', encoding='utf-8') as fs:
                tag_script = soup.script
                tag_script.string = "var testData = " + str(self.testAll)
                fs.write(soup.prettify())
                fs.close()
        except Exception as e:
            raise e


class TestAppender:
    def __init__(self, test_name, feature_name, reporter, client=None):
        self.test_name = test_name
        self.feature_name = feature_name
        self.start_time = None  # Please input the start time by yourself
        self.datas = {}
        self.status = ""
        self.time_takes = ""
        self.steps = []
        self.reporter = reporter
        self.test = {}
        if client is None:
            self.client = "N/A"
        else:
            self.client = client
        self.api_details = []

    def appender_step(self, step_status="", step_description="", use_data="", find_by="",
                      step_details="", screenshot_name=""):
        """
        :param step_description: Optional field -- To log the description for current step
        :param step_status: Required field -- To log current step status
        :param use_data: Optional field -- To log the test data for current step
        :param find_by: Optional field -- To log the object locator for current step
        :param step_details: Optional field -- To log the step logs for current step
        :param screenshot_name: Optional field -- To log the screenshot name for current step
        """
        if not step_status:
            step_status = "failed"
        if not step_description:
            step_description = "N/A"
        if not use_data:
            use_data = "N/A"
        if not find_by:
            find_by = "N/A"
        if not step_details:
            step_details = "N/A"
        wrap_step = {"stepDescription": step_description,
                     "stepStatus": step_status,
                     "data": use_data,
                     "findBy": find_by,
                     "stepDetails": step_details,
                     "screenShot": screenshot_name}
        self.steps.append(wrap_step)

    def completed(self):
        time_format = '%Y-%m-%d %H:%M:%S'
        try:
            # To log the end time when test was completed
            end_time = datetime.now().strftime(time_format)
            self.time_takes = str(
                datetime.strptime(end_time, time_format) - datetime.strptime(
                    self.start_time, time_format))
            # To set the case status if the case steps contains failed or passed.
            self._set_case_status()
            # Build the test details to object test
            self.test = {
                "testName": self.test_name,
                "featureName": self.feature_name,
                "status": self.status,
                "datas": self.datas,
                "timeTakes": self.time_takes,
                "steps": self.steps}
            if self.client is not None:
                self.test["client"] = self.client
            if len(self.api_details) > 0:
                self.test["details"] = self.api_details
            # Add the test details to the reporter
            self.reporter.add_test_details(self.test)
        except Exception as e:
            raise e

    def set_api_details(self, details: list):
        """
         :param details: Optional. To log all request details for each
        """
        temp = []
        if type(details) == list:
            for d in details:
                keys_list = d.keys()
                if "body" not in keys_list:
                    d["body"] = "N/A"
                if "header" not in keys_list:
                    d["header"] = "N/A"
                if "url" not in keys_list:
                    d["url"] = "N/A"
                if "method" not in keys_list:
                    d["method"] = "N/A"
                if "responseBody" not in keys_list:
                    d["responseBody"] = "N/A"
                if "responseHeader" not in keys_list:
                    d["responseHeader"] = "N/A"
                temp.append(d)
            self.api_details = temp
        else:
            self.api_details = []

    def set_test_data(self, datas: dict):
        """
        :param datas: Optional. To log all the datas were used for the current case
        """
        if type(datas) == dict:
            self.datas = datas
        else:
            # Set test data unSuccess
            self.datas = {}
            print(
                "Your input datas type is not correct. Case data set will set to {}. "
                "Please use dict by next time")

    def _set_case_status(self):
        if self.steps:
            for step in self.steps:
                if step['stepStatus'] == "failed":
                    self.status = "failed"
                    break
                self.status = "passed"
        else:
            self.status = "failed"


# Create summary excel report after modify html
# Only display the summary
# If you want to see the test details.
# Please fix comment on self.ws_test_details and self._write_details()
class ExcelCreator:
    def __init__(self, json_path, out_put_path):
        self.json_path = json_path
        self.out_put_path = out_put_path
        self.wb = Workbook()
        try:
            f = open(json_path, 'r')
            self.data = json.load(f)
            f.close()
        except Exception:
            raise Exception(
                f"Please make sure your test json file exist in path {out_put_path}")
        self._create_test_sheet()
        self.style = self._set_style()

    def _create_test_sheet(self):
        self.ws_summary = self.wb.create_sheet("DashBoard", 0)
        # self.ws_test_details = self.wb.create_sheet("TestDetails", 1)

    def _set_style(self):
        my_style = NamedStyle(name="my_style")
        my_style.font = Font(name='Calibri')
        bd = Side(style='thin')
        my_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.wb.add_named_style(my_style)
        return "my_style"

    # Write the summary details in Sheet: DashBoard
    def _write_summary(self):
        # write config summary to 1 - 2
        self._single_line_dict(self.ws_summary, 1, self.data['testConfig'])
        # write feature summary from row 4
        self._multiple_lines_list(self.ws_summary, 4, self.data['featureSummary'])
        # write test details summary from current max row + 2
        from_row = self.ws_summary.max_row + 2
        self._multiple_lines_list(self.ws_summary, from_row, self.data['testDetails'],
                                  ignore_text="steps")
        # write cases total from current max row + 2
        from_row = self.ws_summary.max_row + 2
        self._single_line_dict(self.ws_summary, from_row, self.data['casesChart'])
        # write steps total from current max row + 2
        from_row = self.ws_summary.max_row + 2
        self._single_line_dict(self.ws_summary, from_row, self.data['stepsChart'])
        self._auto_adjust_width(self.ws_summary)

    # Write the test details in Sheet: TestDetails
    # Suggest you to see the test details in html report
    # This method is disable default
    def _write_details(self):
        tests = self.data['testDetails']
        for i in range(len(tests)):
            steps = tests[i]['steps']
            test = tests[i]
            del test['steps']
            if i == 0:
                self._single_line_dict(self.ws_test_details, 1, test)
            else:
                from_row = self.ws_test_details.max_row + 2
                self._single_line_dict(self.ws_test_details, from_row, test)
            from_row = self.ws_test_details.max_row + 1
            self._multiple_lines_list(self.ws_test_details, from_row, steps)
        self._auto_adjust_width(self.ws_test_details)

    def _single_line_dict(self, sheet, from_row, dict_data):
        index = 1
        for key, value in dict_data.items():
            sheet.cell(row=from_row, column=index, value=key.title())
            sheet.cell(row=from_row, column=index).style = self.style
            sheet.cell(row=from_row, column=index).fill = PatternFill('solid',
                                                                      fgColor="95B3D7")
            sheet.cell(row=from_row + 1, column=index, value=str(value))
            sheet.cell(row=from_row + 1, column=index).style = self.style
            index += 1

    def _multiple_lines_list(self, sheet, from_row, list_data, ignore_text=None):
        index = 1
        for k in list_data[0].keys():
            if ignore_text and ignore_text == k:
                continue
            sheet.cell(row=from_row, column=index, value=k.title())
            sheet.cell(row=from_row, column=index).style = self.style
            sheet.cell(row=from_row, column=index).fill = PatternFill('solid',
                                                                      fgColor="95B3D7")
            index += 1
        for index in range(len(list_data)):
            dict_data = list_data[index]
            v_index = 1
            row = from_row + index + 1
            for k, v in dict_data.items():
                if ignore_text and ignore_text == k:
                    continue
                v = str(v)
                sheet.cell(row=row, column=v_index, value=v)
                sheet.cell(row=row, column=v_index).style = self.style
                v_index += 1

    def _auto_adjust_width(self, sheet):
        # auto adjust the width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            sheet.column_dimensions[column].width = max_length + 2

    def save(self):
        try:
            self._write_summary()
            # self._write_details()
            self.wb.save(self.out_put_path)
        except Exception as e:
            raise e
