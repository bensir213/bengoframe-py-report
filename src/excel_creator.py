from openpyxl import Workbook
import os
from datetime import datetime
from openpyxl.styles import NamedStyle, Font, Side, Border, Color, PatternFill
import json


# Create summary excel report after modify html
# Only display the summary
# If you want to see the test details. Please fix comment on self.ws_test_details and self._write_details()
class ExcelCreator:
    def __init__(self, json_path, out_put_path):
        self.json_path = json_path
        self.out_put_path = out_put_path
        self.wb = Workbook()
        try:
            f = open(json_path, 'r')
            self.data = json.load(f)
            f.close()
        except Exception:
            raise Exception(f"Please make sure your test json file exist in path {out_put_path}")
        self._create_test_sheet()
        self.style = self._set_style()

    def _create_test_sheet(self):
        self.ws_summary = self.wb.create_sheet("DashBoard", 0)
        # self.ws_test_details = self.wb.create_sheet("TestDetails", 1)

    def _set_style(self):
        my_style = NamedStyle(name="my_style")
        my_style.font = Font(name='Calibri')
        bd = Side(style='thin')
        my_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.wb.add_named_style(my_style)
        return "my_style"

    # Write the summary details in Sheet: DashBoard
    def _write_summary(self):
        # write config summary to 1 - 2
        self._single_line_dict(self.ws_summary, 1, self.data['testConfig'])
        # write feature summary from row 4
        self._multiple_lines_list(self.ws_summary, 4, self.data['featureSummary'])
        # write test details summary from current max row + 2
        from_row = self.ws_summary.max_row + 2
        self._multiple_lines_list(self.ws_summary, from_row, self.data['testDetails'], ignore_text="steps")
        # write cases total from current max row + 2
        from_row = self.ws_summary.max_row + 2
        self._single_line_dict(self.ws_summary, from_row, self.data['casesChart'])
        # write steps total from current max row + 2
        from_row = self.ws_summary.max_row + 2
        self._single_line_dict(self.ws_summary, from_row, self.data['stepsChart'])
        self._auto_adjust_width(self.ws_summary)

    # Write the test details in Sheet: TestDetails
    # Suggest you to see the test details in html report
    # This method is disable default
    def _write_details(self):
        tests = self.data['testDetails']
        for i in range(len(tests)):
            steps = tests[i]['steps']
            test = tests[i]
            del test['steps']
            if i == 0:
                self._single_line_dict(self.ws_test_details, 1, test)
            else:
                from_row = self.ws_test_details.max_row + 2
                self._single_line_dict(self.ws_test_details, from_row, test)
            from_row = self.ws_test_details.max_row + 1
            self._multiple_lines_list(self.ws_test_details, from_row, steps)
        self._auto_adjust_width(self.ws_test_details)

    def _single_line_dict(self, sheet, from_row, dict_data):
        index = 1
        for key, value in dict_data.items():
            sheet.cell(row=from_row, column=index, value=key.title())
            sheet.cell(row=from_row, column=index).style = self.style
            sheet.cell(row=from_row, column=index).fill = PatternFill('solid', fgColor="95B3D7")
            sheet.cell(row=from_row + 1, column=index, value=str(value))
            sheet.cell(row=from_row + 1, column=index).style = self.style
            index += 1

    def _multiple_lines_list(self, sheet, from_row, list_data, ignore_text=None):
        index = 1
        for k in list_data[0].keys():
            if ignore_text and ignore_text == k:
                continue
            sheet.cell(row=from_row, column=index, value=k.title())
            sheet.cell(row=from_row, column=index).style = self.style
            sheet.cell(row=from_row, column=index).fill = PatternFill('solid', fgColor="95B3D7")
            index += 1
        for index in range(len(list_data)):
            dict_data = list_data[index]
            v_index = 1
            row = from_row + index + 1
            for k, v in dict_data.items():
                if ignore_text and ignore_text == k:
                    continue
                v = str(v)
                sheet.cell(row=row, column=v_index, value=v)
                sheet.cell(row=row, column=v_index).style = self.style
                v_index += 1

    def _auto_adjust_width(self, sheet):
        # auto adjust the width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            sheet.column_dimensions[column].width = max_length + 2

    def save(self):
        try:
            self._write_summary()
            # self._write_details()
            self.wb.save(self.out_put_path)
        except Exception as e:
            raise e
